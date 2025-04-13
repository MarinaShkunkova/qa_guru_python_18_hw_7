from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from tests.conftest import ARCHIVE_FILE_PATH


def test_pdf_reader():
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zip_file:
        with zip_file.open("file_for_testing_pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            assert "Здравствуйте!" in reader.pages[0].extract_text()


def test_xlsx_reader():
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zip_file:
        with zip_file.open("file_for_testing_xlxs.xlsx") as xlsx_file:
            reader = load_workbook(xlsx_file)
            assert "Postcode" == reader.active.cell(row=1, column=1).value


def test_csv_reader():
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zip_file:
        with zip_file.open("file_for_testing_csv.csv") as csv_file:
            reader = csv_file.read().decode("utf-8")
            assert "Alice" in reader
